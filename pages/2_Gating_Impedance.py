import io
import os
import tempfile

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st

from core.gating import apply_time_gate, compute_tdr_gated, gate_tdr_csv
from core.impedance import compute_tdr
from core.mixed_mode import single_to_mixed_mode, single_to_mixed_mode_npairs
from core.parser import get_frequency_ghz, get_smatrix, load_s4p, load_s16p

st.set_page_config(page_title="Gating Impedance", layout="wide")
st.title("SI Tool — Gating Impedance")

_PAIR_OPTS = [
    "Pair 1 (Diff 1 ↔ Diff 2)",
    "Pair 2 (Diff 3 ↔ Diff 4)",
    "Pair 3 (Diff 5 ↔ Diff 6)",
    "Pair 4 (Diff 7 ↔ Diff 8)",
]

_AX = dict(showgrid=True, gridcolor="#888888", gridwidth=1, showline=False,
           mirror=False, zeroline=False,
           title_font=dict(size=20, family="Arial"), tickfont=dict(size=16, family="Arial"))
_LY = dict(hovermode="x unified", font=dict(size=16, family="Arial"),
           title_font=dict(size=22, family="Arial"),
           legend=dict(orientation="v", yanchor="bottom", y=0.02, xanchor="right", x=0.98,
                       font=dict(size=16, family="Arial")),
           margin=dict(t=42, b=58, l=60, r=20),
           plot_bgcolor="#ffffff", paper_bgcolor="#ffffff")
_BL = dict(color="#888888", width=1)
_EPS = 1e-9


def _patch_chart_sppr(xlsx_bytes: bytes) -> bytes:
    """Post-process xlsx: inject chart area spPr + plot area manualLayout into each chart."""
    import zipfile
    from lxml import etree

    NS_C = 'http://schemas.openxmlformats.org/drawingml/2006/chart'
    NS_A = 'http://schemas.openxmlformats.org/drawingml/2006/main'

    def _chart_area_sppr():
        spPr = etree.Element(f'{{{NS_C}}}spPr')
        sf = etree.SubElement(spPr, f'{{{NS_A}}}solidFill')
        etree.SubElement(sf, f'{{{NS_A}}}schemeClr').set('val', 'bg1')
        ln = etree.SubElement(spPr, f'{{{NS_A}}}ln')
        ln.set('w', '9525'); ln.set('cap', 'flat'); ln.set('cmpd', 'sng'); ln.set('algn', 'ctr')
        etree.SubElement(ln, f'{{{NS_A}}}noFill')
        etree.SubElement(ln, f'{{{NS_A}}}round')
        etree.SubElement(spPr, f'{{{NS_A}}}effectLst')
        return spPr

    def _plot_area_layout():
        layout = etree.Element(f'{{{NS_C}}}layout')
        ml = etree.SubElement(layout, f'{{{NS_C}}}manualLayout')
        for tag, val in [('layoutTarget', 'inner'), ('xMode', 'edge'), ('yMode', 'edge'),
                         ('x', '0.10992255198869372'), ('y', '0.10568893876292018'),
                         ('w', '0.83791986001749785'), ('h', '0.75324590629372912')]:
            etree.SubElement(ml, f'{{{NS_C}}}{tag}').set('val', val)
        return layout

    buf_in  = io.BytesIO(xlsx_bytes)
    buf_out = io.BytesIO()
    with zipfile.ZipFile(buf_in, 'r') as zin, \
         zipfile.ZipFile(buf_out, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.startswith('xl/charts/chart') and item.filename.endswith('.xml'):
                root = etree.fromstring(data)
                # chart area spPr
                if root.find(f'{{{NS_C}}}spPr') is None:
                    root.append(_chart_area_sppr())
                # plot area manualLayout
                pa = root.find(f'.//{{{NS_C}}}plotArea')
                if pa is not None and pa.find(f'{{{NS_C}}}layout') is None:
                    pa.insert(0, _plot_area_layout())
                data = etree.tostring(root, xml_declaration=True,
                                      encoding='UTF-8', standalone=True)
            zout.writestr(item, data)
    buf_out.seek(0)
    return buf_out.read()


def _build_xlsx(rl_freq, rl_orig, rl_gated, imp_time, imp_orig, imp_gated,
                z_min=70, z_max=120, rl_min=-80) -> bytes:
    from openpyxl import Workbook
    from openpyxl.chart import ScatterChart, Reference, Series
    from openpyxl.chart.axis import ChartLines
    from openpyxl.chart.data_source import NumFmt
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.chart.text import RichText, Text
    from openpyxl.chart.title import Title
    from openpyxl.drawing.colors import ColorChoice, SchemeColor
    from openpyxl.drawing.line import LineProperties
    from openpyxl.drawing.text import (CharacterProperties, Font as DFont,
                                        Paragraph, ParagraphProperties,
                                        RegularTextRun, RichTextProperties)

    def _rich_title(text, sz, rot=0):
        f = DFont(typeface='Arial')
        rpr_def = CharacterProperties(sz=sz, b=False, strike='noStrike',
                                       kern=1200, spc=0, baseline=0, latin=f)
        rpr_run = CharacterProperties(sz=sz, baseline=0, latin=f)
        para = Paragraph(
            pPr=ParagraphProperties(defRPr=rpr_def),
            r=[RegularTextRun(rPr=rpr_run, t=text)],
        )
        body = RichTextProperties(rot=rot, spcFirstLastPara=True,
                                   vertOverflow='ellipsis', wrap='square',
                                   anchor='ctr', anchorCtr=True)
        return Title(tx=Text(rich=RichText(bodyPr=body, p=[para])))

    def _scheme_fill(val, lumMod=None, lumOff=None):
        sc = SchemeColor(val=val)
        if lumMod: sc.lumMod = lumMod
        if lumOff: sc.lumOff = lumOff
        cc = ColorChoice(); cc.schemeClr = sc
        return cc

    def _gridlines(w, val, lumMod=None, lumOff=None):
        ln = LineProperties(w=w, solidFill=_scheme_fill(val, lumMod, lumOff))
        gl = ChartLines(); gl.spPr = GraphicalProperties(ln=ln)
        return gl

    def _make_series(ws, x_col, y_col, n, scheme_color):
        xref = Reference(ws, min_col=x_col, min_row=2,     max_row=n + 1)
        yref = Reference(ws, min_col=y_col, min_row=1,     max_row=n + 1)
        s = Series(yref, xref, title_from_data=True)
        s.smooth = True
        s.graphicalProperties.line.width = 28575
        s.graphicalProperties.line.solidFill = _scheme_fill(scheme_color)
        s.marker.symbol = "none"
        return s

    def _make_chart(title_text, x_title, y_title, y_min, y_max):
        ch = ScatterChart()
        ch.scatterStyle    = "smoothMarker"
        ch.width           = 15
        ch.height          = 7.5
        ch.title           = _rich_title(title_text, sz=2400)
        ch.legend.position = "r"
        ch.roundedCorners = False
        # plot area: white fill, no border
        pa_spPr = GraphicalProperties(solidFill="FFFFFF")
        pa_spPr.ln = LineProperties(noFill=True)
        ch.plot_area.spPr = pa_spPr
        # x axis
        ch.x_axis.title           = _rich_title(x_title, sz=2000)
        ch.x_axis.crossBetween    = "midCat"
        ch.x_axis.numFmt          = NumFmt(formatCode='General', sourceLinked=True)
        ch.x_axis.delete          = False
        ch.x_axis.majorGridlines  = _gridlines(9525, 'tx1', lumMod=15000, lumOff=85000)
        # y axis
        ch.y_axis.title           = _rich_title(y_title, sz=2000, rot=-5400000)
        ch.y_axis.crossBetween    = "midCat"
        ch.y_axis.numFmt          = NumFmt(formatCode='General', sourceLinked=True)
        ch.y_axis.delete          = False
        ch.y_axis.majorGridlines  = _gridlines(6350, 'bg2', lumMod=75000)
        ch.y_axis.scaling.min     = float(y_min)
        ch.y_axis.scaling.max     = float(y_max)
        return ch

    wb = Workbook()
    wb.remove(wb.active)

    # ── Return Loss ──────────────────────────────────────────────────────────
    ws_rl = wb.create_sheet("Return_Loss")
    ws_rl.append(['Frequency [GHz]', 'RL_Original', 'RL_Gated'])
    for f, ro, rg in zip(rl_freq, rl_orig, rl_gated):
        ws_rl.append([round(float(f), 6), round(float(ro), 4), round(float(rg), 4)])
    n_rl = len(rl_freq)
    ch_rl = _make_chart("Return Loss", "Frequency [GHz]", "Magnitude [dB]", rl_min, 0)
    ch_rl.series.append(_make_series(ws_rl, 1, 2, n_rl, 'accent1'))
    ch_rl.series.append(_make_series(ws_rl, 1, 3, n_rl, 'accent2'))
    wb.create_chartsheet("Chart1").add_chart(ch_rl)

    # ── Impedance ────────────────────────────────────────────────────────────
    ws_imp = wb.create_sheet("Impedance")
    ws_imp.append(['Time [ns]', 'Z_Original', 'Z_Gated'])
    for t, zo, zg in zip(imp_time, imp_orig, imp_gated):
        ws_imp.append([round(float(t), 6), round(float(zo), 4), round(float(zg), 4)])
    n_imp = len(imp_time)
    ch_imp = _make_chart("Impedance", "Time [ns]", "Z [Ohms]", z_min, z_max)
    ch_imp.series.append(_make_series(ws_imp, 1, 2, n_imp, 'accent1'))
    ch_imp.series.append(_make_series(ws_imp, 1, 3, n_imp, 'accent2'))
    wb.create_chartsheet("Chart2").add_chart(ch_imp)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return _patch_chart_sppr(buf.read())


def _baseline_correct(t: np.ndarray, z_gated: np.ndarray, z_orig: np.ndarray,
                      t_stop: float, z0: float = 100.0) -> np.ndarray:
    """Shift post-gate gamma so gated impedance aligns with original at t_stop."""
    gamma_orig  = (z_orig  - z0) / (z_orig  + z0 + 1e-15)
    gamma_gated = (z_gated - z0) / (z_gated + z0 + 1e-15)
    idx = int(np.searchsorted(t, t_stop))
    idx = min(idx, len(t) - 1)
    offset = float(gamma_orig[idx] - gamma_gated[idx])
    gamma_c = gamma_gated.copy()
    gamma_c[idx:] += offset
    gamma_c = np.clip(gamma_c, -0.9999, 0.9999)
    return z0 * (1 + gamma_c) / (1 - gamma_c)


# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("設定")

    src = st.radio("資料來源", ["VNA (S4P/S16P)", "TDR CSV"], horizontal=True, key="gi_src")
    vna_mode = src.startswith("VNA")

    # ── VNA mode ──
    if vna_mode:
        ftype = st.radio("檔案類型", ["S4P", "S16P"], horizontal=True, key="gi_ftype")
        uploaded_vna = st.file_uploader(f"上傳 {ftype} 檔案",
                                        type=[ftype.lower()], key="gi_up")
        pair_idx = 0
        if ftype == "S16P":
            sel = st.selectbox("選擇差分對", _PAIR_OPTS, key="gi_pair")
            pair_idx = _PAIR_OPTS.index(sel)

    # ── CSV mode ──
    else:
        uploaded_csv = st.file_uploader("上傳 TDR CSV 檔案", type=["csv"], key="gi_csv")
        z0_csv = st.number_input("參考阻抗 Z0 (Ω)", 1, 300, 100, 1, key="gi_z0")

    st.divider()

    # ── VNA-only settings ──
    if vna_mode:
        rise_ps  = st.number_input("TDR Rise Time (ps)", 10, 200, 35, 5, key="gi_rps")
    x_max_ns = st.number_input("TDR 顯示範圍 (ns)", 0.1, 50.0, 5.0, 0.1,
                                format="%.1f", key="gi_xns")

    st.divider()
    st.subheader("Gate 設定")
    gate_mode = st.radio("模式", ["Exclude（移除選取）", "Include（保留選取）"], key="gi_gm")
    exclude  = gate_mode.startswith("Exclude")
    t_start  = st.number_input("Gate Start (ns)", 0.0, 100.0, 0.0,  0.05, format="%.2f", key="gi_ts")
    t_stop   = st.number_input("Gate Stop  (ns)", 0.0, 100.0, 0.5,  0.05, format="%.2f", key="gi_te")
    rise_pct = st.slider("Gate Rise (%)", 1, 30, 10, key="gi_rp")
    strength = st.slider("Gate 強度 (%)", 0, 100, 100, key="gi_str") / 100.0

    st.divider()
    st.subheader("Y 軸")
    z_min  = st.number_input("Impedance Min (Ω)", 0,   300,  60, 10, key="gi_zn")
    z_max  = st.number_input("Impedance Max (Ω)", 0,   300, 140, 10, key="gi_zx")
    rl_min = st.number_input("RL Min (dB)",      -120,   0, -80, 10, key="gi_rn")

# ── Validation ────────────────────────────────────────────────────────────────
if t_start >= t_stop:
    st.warning("Gate Start 必須小於 Gate Stop")
    st.stop()

x_rng      = [0 - _EPS, float(x_max_ns) + _EPS]
mode_label = "Exclude" if exclude else "Include"
rise_frac  = rise_pct / 100.0

# ══════════════════════════════════════════════════════════════════════════════
# VNA MODE
# ══════════════════════════════════════════════════════════════════════════════
if vna_mode:
    if uploaded_vna is None:
        st.info("← 請從左側上傳 S4P 或 S16P 檔案")
        st.stop()

    try:
        with tempfile.NamedTemporaryFile(suffix=f".{ftype.lower()}", delete=False) as f:
            f.write(uploaded_vna.read())
            tmp = f.name

        if ftype == "S4P":
            nw    = load_s4p(tmp)
            s_mm  = single_to_mixed_mode(get_smatrix(nw), mapping='A')
            sdd11 = s_mm[:, 0, 0]
        else:
            nw    = load_s16p(tmp)
            s_mm  = single_to_mixed_mode_npairs(get_smatrix(nw), n_pairs=4, mapping='A')
            sdd11 = s_mm[:, 4 * pair_idx, 4 * pair_idx]

        os.unlink(tmp)
        freq_ghz = get_frequency_ghz(nw)
        freqs_hz = freq_ghz * 1e9

    except Exception as e:
        st.error(f"檔案讀取失敗：{e}")
        st.stop()

    # Gating
    sdd11_gated, _, _, _ = apply_time_gate(
        sdd11, freqs_hz, t_start, t_stop,
        rise_frac=rise_frac, exclude=exclude, strength=strength,
    )
    t_ps_o, z_o = compute_tdr(sdd11, freqs_hz, z0=100.0, rise_time_ps=float(rise_ps))
    t_ns_o = t_ps_o / 1000

    t_ps_g, z_g, t_ns_g, _ = compute_tdr_gated(
        sdd11, freqs_hz, t_start, t_stop,
        z0=100.0, rise_time_ps=float(rise_ps),
        rise_frac=rise_frac, exclude=exclude, strength=strength,
    )

    z_g = _baseline_correct(t_ns_g, z_g, z_o, t_stop, z0=100.0)

    rl_o = 20 * np.log10(np.abs(sdd11)       + 1e-15)
    rl_g = 20 * np.log10(np.abs(sdd11_gated) + 1e-15)
    f_max = float(freq_ghz[-1])

    # Plot 1: TDR + Gate
    fig1 = go.Figure()
    fig1.add_trace(go.Scatter(x=t_ns_o, y=z_o, name="TDR 原始",  line=dict(color="#2563eb")))
    fig1.add_vrect(x0=t_start, x1=t_stop, fillcolor="#fbbf24",
                   opacity=0.25, layer="below", line_width=0)
    fig1.add_vline(x=t_start, line=dict(color="#d97706", dash="dash", width=1.5))
    fig1.add_vline(x=t_stop,  line=dict(color="#d97706", dash="dash", width=1.5))
    fig1.add_hline(y=z_min, line=_BL); fig1.add_hline(y=z_max, line=_BL)
    fig1.add_vline(x=0, line=_BL);     fig1.add_vline(x=x_max_ns, line=_BL)
    fig1.update_layout(
        title=dict(text=f"TDR — Gate 範圍 ({t_start:.2f} ~ {t_stop:.2f} ns, {mode_label})",
                   x=0.5, xanchor="center", pad=dict(t=20)),
        xaxis_title="Time (ns)", yaxis_title="Impedance (Ω)",
        xaxis=dict(**_AX, range=x_rng),
        yaxis=dict(**_AX, range=[z_min - _EPS, z_max + _EPS], tick0=z_min),
        **_LY,
    )
    st.plotly_chart(fig1, use_container_width=True)

    # Plot 2: Return Loss
    fig2 = go.Figure()
    fig2.add_trace(go.Scatter(x=freq_ghz, y=rl_o, name="SDD11 原始",          line=dict(color="#2563eb")))
    fig2.add_trace(go.Scatter(x=freq_ghz, y=rl_g, name=f"SDD11 Gated ({mode_label})", line=dict(color="#dc2626")))
    fig2.add_hline(y=0,      line=_BL); fig2.add_hline(y=rl_min, line=_BL)
    fig2.add_vline(x=0,      line=_BL); fig2.add_vline(x=f_max,  line=_BL)
    fig2.update_layout(
        title=dict(text="Return Loss — Gating 前後對比", x=0.5, xanchor="center", pad=dict(t=20)),
        xaxis_title="Frequency (GHz)", yaxis_title="Magnitude (dB)",
        xaxis=dict(**_AX, range=[0 - _EPS, f_max + _EPS]),
        yaxis=dict(**_AX, range=[rl_min - _EPS, 0 + _EPS], tick0=0),
        **_LY,
    )
    st.plotly_chart(fig2, use_container_width=True)

    # Plot 3: Impedance before/after
    fig3 = go.Figure()
    fig3.add_trace(go.Scatter(x=t_ns_o, y=z_o, name="Impedance 原始",          line=dict(color="#2563eb")))
    fig3.add_trace(go.Scatter(x=t_ns_g, y=z_g, name=f"Impedance Gated ({mode_label})", line=dict(color="#dc2626")))
    fig3.add_vline(x=t_start, line=dict(color="#d97706", dash="dash", width=1.5))
    fig3.add_vline(x=t_stop,  line=dict(color="#d97706", dash="dash", width=1.5))
    fig3.add_hline(y=z_min, line=_BL); fig3.add_hline(y=z_max, line=_BL)
    fig3.add_vline(x=0, line=_BL);     fig3.add_vline(x=x_max_ns, line=_BL)
    fig3.update_layout(
        title=dict(text="Impedance — Gating 前後對比", x=0.5, xanchor="center", pad=dict(t=20)),
        xaxis_title="Time (ns)", yaxis_title="Impedance (Ω)",
        xaxis=dict(**_AX, range=x_rng),
        yaxis=dict(**_AX, range=[z_min - _EPS, z_max + _EPS], tick0=z_min),
        **_LY,
    )
    st.plotly_chart(fig3, use_container_width=True)

    with st.sidebar:
        st.divider()
        st.subheader("輸出")
        xlsx_data = _build_xlsx(freq_ghz, rl_o, rl_g, t_ns_o, z_o, z_g, z_min, z_max, rl_min)
        st.download_button("下載 Excel", xlsx_data, "Gating_Export.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)

# ══════════════════════════════════════════════════════════════════════════════
# TDR CSV MODE
# ══════════════════════════════════════════════════════════════════════════════
else:
    if uploaded_csv is None:
        st.info("← 請從左側上傳 TDR CSV 檔案")
        st.stop()

    try:
        df_csv   = pd.read_csv(uploaded_csv)
        t_ns_csv = df_csv.iloc[:, 0].to_numpy(dtype=float)
        z_csv    = df_csv.iloc[:, 1].to_numpy(dtype=float)
    except Exception as e:
        st.error(f"CSV 讀取失敗：{e}")
        st.stop()

    # Gating
    z_gated, s11_orig, s11_gated, freq_ghz_csv, _ = gate_tdr_csv(
        t_ns_csv, z_csv, t_start, t_stop,
        z0=float(z0_csv), rise_frac=rise_frac,
        exclude=exclude, strength=strength,
    )

    z_gated = _baseline_correct(t_ns_csv, z_gated, z_csv, t_stop, z0=float(z0_csv))

    rl_o = 20 * np.log10(np.abs(s11_orig)  + 1e-15)
    rl_g = 20 * np.log10(np.abs(s11_gated) + 1e-15)

    # Plot 1: TDR + Gate
    fig1 = go.Figure()
    fig1.add_trace(go.Scatter(x=t_ns_csv, y=z_csv, name="TDR 原始", line=dict(color="#2563eb")))
    fig1.add_vrect(x0=t_start, x1=t_stop, fillcolor="#fbbf24",
                   opacity=0.25, layer="below", line_width=0)
    fig1.add_vline(x=t_start, line=dict(color="#d97706", dash="dash", width=1.5))
    fig1.add_vline(x=t_stop,  line=dict(color="#d97706", dash="dash", width=1.5))
    fig1.add_hline(y=z_min, line=_BL); fig1.add_hline(y=z_max, line=_BL)
    fig1.add_vline(x=0, line=_BL);     fig1.add_vline(x=x_max_ns, line=_BL)
    fig1.update_layout(
        title=dict(text=f"TDR — Gate 範圍 ({t_start:.2f} ~ {t_stop:.2f} ns, {mode_label})",
                   x=0.5, xanchor="center", pad=dict(t=20)),
        xaxis_title="Time (ns)", yaxis_title="Impedance (Ω)",
        xaxis=dict(**_AX, range=x_rng),
        yaxis=dict(**_AX, range=[z_min - _EPS, z_max + _EPS], tick0=z_min),
        **_LY,
    )
    st.plotly_chart(fig1, use_container_width=True)

    # Plot 2: Return Loss (from FFT of impulse)
    f_max_csv = float(freq_ghz_csv[-1])
    fig2 = go.Figure()
    fig2.add_trace(go.Scatter(x=freq_ghz_csv, y=rl_o, name="RL 原始",          line=dict(color="#2563eb")))
    fig2.add_trace(go.Scatter(x=freq_ghz_csv, y=rl_g, name=f"RL Gated ({mode_label})", line=dict(color="#dc2626")))
    fig2.add_hline(y=0,      line=_BL); fig2.add_hline(y=rl_min,   line=_BL)
    fig2.add_vline(x=0,      line=_BL); fig2.add_vline(x=f_max_csv, line=_BL)
    fig2.update_layout(
        title=dict(text="Return Loss — Gating 前後對比（由 TDR 推算）",
                   x=0.5, xanchor="center", pad=dict(t=20)),
        xaxis_title="Frequency (GHz)", yaxis_title="Magnitude (dB)",
        xaxis=dict(**_AX, range=[0 - _EPS, f_max_csv + _EPS]),
        yaxis=dict(**_AX, range=[rl_min - _EPS, 0 + _EPS], tick0=0),
        **_LY,
    )
    st.plotly_chart(fig2, use_container_width=True)

    # Plot 3: Impedance before/after
    fig3 = go.Figure()
    fig3.add_trace(go.Scatter(x=t_ns_csv, y=z_csv,   name="Impedance 原始",          line=dict(color="#2563eb")))
    fig3.add_trace(go.Scatter(x=t_ns_csv, y=z_gated, name=f"Impedance Gated ({mode_label})", line=dict(color="#dc2626")))
    fig3.add_vline(x=t_start, line=dict(color="#d97706", dash="dash", width=1.5))
    fig3.add_vline(x=t_stop,  line=dict(color="#d97706", dash="dash", width=1.5))
    fig3.add_hline(y=z_min, line=_BL); fig3.add_hline(y=z_max, line=_BL)
    fig3.add_vline(x=0, line=_BL);     fig3.add_vline(x=x_max_ns, line=_BL)
    fig3.update_layout(
        title=dict(text="Impedance — Gating 前後對比", x=0.5, xanchor="center", pad=dict(t=20)),
        xaxis_title="Time (ns)", yaxis_title="Impedance (Ω)",
        xaxis=dict(**_AX, range=x_rng),
        yaxis=dict(**_AX, range=[z_min - _EPS, z_max + _EPS], tick0=z_min),
        **_LY,
    )
    st.plotly_chart(fig3, use_container_width=True)

    with st.sidebar:
        st.divider()
        st.subheader("輸出")
        xlsx_data = _build_xlsx(freq_ghz_csv, rl_o, rl_g, t_ns_csv, z_csv, z_gated, z_min, z_max, rl_min)
        st.download_button("下載 Excel", xlsx_data, "Gating_Export.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)
